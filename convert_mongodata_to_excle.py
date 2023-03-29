from openpyxl.reader.excel import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from pymongo import MongoClient
import certifi
import pandas as pd
from openpyxl import Workbook 
from openpyxl.styles import PatternFill, Alignment, Side, Border




client = MongoClient('mongodb+srv://19l31a0581:fenA5B7Qr9FtFjw5@cluster0.9mhf5ll.mongodb.net/test', tlsCAFile=certifi.where())
db = client['contestDetails']
collection = db['ranklists']

def ranklist_by_year():
    # Load the Excel workbook
    divs = ['Div A','Div B','Div C','Div D']
    years = ['I Year','II Year','III Year','IV Year']

    for year in years:
        first_year=[]
        second_year=[]
        third_year=[]
        fourth_year=[]
        wb = load_workbook('ranklist.xlsx')
        for div in divs:
            # Select the worksheet you want to read
            ws = wb[f'{div}']

            # Iterate through the rows
            for row in ws.iter_rows(min_row=1, values_only=True):
                if(len(row)<7): break
                if(row[7]=='I Year'): 
                    temp_row = tuple(list(row) + [f'{div}'])
                    first_year.append(temp_row)
                if(row[7]=='II Year'): 
                    temp_row = tuple(list(row) + [f'{div}'])
                    second_year.append(temp_row)
                if(row[7]=='III Year'): 
                    temp_row = tuple(list(row) + [f'{div}'])
                    third_year.append(temp_row)
                if(row[7]=='IV Year'): 
                    temp_row = tuple(list(row) + [f'{div}'])
                    fourth_year.append(temp_row)
        try:
            wb = load_workbook('Yearwise_ranklist.xlsx')
        except:
            wb = Workbook()
            wb.save('Yearwise_ranklist.xlsx')
        try:
            del wb['Sheet']
        except:
            pass
        try:
            del wb[year]
            wb.create_sheet(year)
        except:
            wb.create_sheet(year)
        ws = wb[year]
        print(year)
        col_tup=('Rank','UserId','Total Score','Total Time','Name','Roll number','Branch','Year','Section','Profile url','Division')
        if(year == 'I Year') : 
            first_year.insert(0,col_tup)
            df = pd.DataFrame(first_year)
        if(year == 'II Year') :
            second_year.insert(0,col_tup)
            df = pd.DataFrame(second_year)
        if(year == 'III Year') : 
            third_year.insert(0,col_tup)
            df = pd.DataFrame(third_year)
        if(year == 'IV Year') : 
            fourth_year.insert(0,col_tup)
            df = pd.DataFrame(fourth_year)
        print(len(df))
        if len(df)==1:
            ws.column_dimensions['A'].width = 30
            ws['A1']=f'No participants in {year}'
        else:
            for r in dataframe_to_rows(df, index=False, header=False):
                # print("r=",r)
                ws.append(r)

            # Save the workbook to a file
            for col in ws.columns: 
                for cell in col:
                    alignment_obj = cell.alignment.copy(horizontal='center', vertical='center')
                    cell.alignment = alignment_obj
            double = Side(border_style="double", color="000000")

            col = ['A','B','C','D','E','F','G','H','I','J','K']
            for column in col:
                ws[f'{column}1'].fill = PatternFill('solid', fgColor='FFFF00')
                ws[f'{column}1'].alignment = Alignment(horizontal='center')
                ws[f'{column}1'].border = Border(top=double, left=double, right=double, bottom=double)
                if column!='J' and column!='E': ws.column_dimensions[f'{column}'].width = 20
                else: ws.column_dimensions[f'{column}'].width = 40
        wb.save('Yearwise_ranklist.xlsx')


def convert_to_excel(contest_id):
    # Retrieve all data from the collection
    divisions = ['Div A','Div B','Div C','Div D']
    for div in divisions:
        data = list(collection.aggregate([
            {'$match':
                 {
                     'contest': contest_id,
                     'division': div
                 }
            },
            {
                '$sort' : {'Rank' : 1}
            },
            {
                '$project' :
                    {
                        '_id' :0,
                        'Rank' : 1,
                        'Username' : 1,
                        'Total score' : 1,
                        'Total time' : 1
                    }
            }
        ]))
        # print("data=",data)
        userCollection = db['codechef_users']
        for user in data : 
            username = user['Username']
            rollnumber = '-'
            name = '-'
            branch = '-'
            section = '-'
            year = '-'
            codechefUrl = '-'
            try:
                userdata = list(userCollection.find({'userId' : username}))[0]
                name = userdata['name']
                rollnumber = userdata['rollNo']
                branch = userdata['branch']
                section = userdata['section']
                year = userdata['year']
                codechefUrl = userdata['url']
            except:
                pass
            user['Name'] = name
            user['Roll number'] = rollnumber
            user['Branch'] = branch
            user['year'] = year
            user['section'] = section
            user['codechefUrl'] = codechefUrl
        # print('div=',div)
        # print("total_cout=",data)
        # Convert the data to a DataFrame
        df = pd.DataFrame(data)

        # Create a new Excel workbook and add the DataFrame to a new worksheet
        wb = Workbook()
        try:
            wb = load_workbook('ranklist.xlsx')
        except:
            wb = Workbook()
            wb.save('ranklist.xlsx')
        try:
            del wb['Sheet']
        except:
            pass
        try:
            del wb[div]
            wb.create_sheet(div)
        except:
            wb.create_sheet(div)
        ws = wb[div]

        if df.empty:
            ws.column_dimensions['A'].width = 30
            ws['A1']=f'No participants in {div}'
        else:
            for r in dataframe_to_rows(df, index=False, header=True):
                # print("r=",r)
                ws.append(r)

            # Save the workbook to a file
            for col in ws.columns: 
                for cell in col:
                    alignment_obj = cell.alignment.copy(horizontal='center', vertical='center')
                    cell.alignment = alignment_obj
            double = Side(border_style="double", color="000000")

            col = ['A','B','C','D','E','F','G','H','I','J']
            for column in col:
                ws[f'{column}1'].fill = PatternFill('solid', fgColor='FFFF00')
                ws[f'{column}1'].alignment = Alignment(horizontal='center')
                ws[f'{column}1'].border = Border(top=double, left=double, right=double, bottom=double)
                if column!='J' and column!='E': ws.column_dimensions[f'{column}'].width = 20
                else: ws.column_dimensions[f'{column}'].width = 40
            # ws['A1'].fill = PatternFill('solid', fgColor='FFFF00')
            # ws['A1'].alignment = Alignment(horizontal='center')
            # ws['A1'].border = Border(top=double, left=double, right=double, bottom=double)
            # ws.column_dimensions['A'].width = 20

            # ws['B1'].fill = PatternFill('solid', fgColor='FFFF00')
            # ws['B1'].alignment = Alignment(horizontal='center')
            # ws['B1'].border = Border(top=double, left=double, right=double, bottom=double)
            # ws.column_dimensions['B'].width = 20

            # ws['C1'].fill = PatternFill('solid', fgColor='FFFF00')
            # ws['C1'].alignment = Alignment(horizontal='center')
            # ws['C1'].border = Border(top=double, left=double, right=double, bottom=double)
            # ws.column_dimensions['C'].width = 20

            # ws['D1'].fill = PatternFill('solid', fgColor='FFFF00')
            # ws['D1'].alignment = Alignment(horizontal='center')
            # ws['D1'].border = Border(top=double, left=double, right=double, bottom=double)
            # ws.column_dimensions['D'].width = 20

            # ws['E1'].fill = PatternFill('solid', fgColor='FFFF00')
            # ws['E1'].alignment = Alignment(horizontal='center')
            # ws['E1'].border = Border(top=double, left=double, right=double, bottom=double)
            # ws.column_dimensions['E'].width = 20
            
        wb.save('ranklist.xlsx')


def check_if_contest_exists_in_db(contest_id):
    data = list(collection.aggregate([
        {
            '$match' : {
                'contest' : contest_id
            }
        }
    ]))
    if(len(data)!=0):
        return 1
    return 0

# convert_to_excel("START78")

# ranklist_by_year();