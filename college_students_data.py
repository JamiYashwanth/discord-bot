from concurrent.futures import ThreadPoolExecutor, as_completed
from selenium import webdriver
from pymongo import MongoClient, ASCENDING, errors
from selenium.webdriver.common.keys import Keys
import time
from openpyxl import Workbook, load_workbook
from bs4 import BeautifulSoup
import certifi
# from test import *
from convert_mongodata_to_excle import *
client = MongoClient('mongodb+srv://19l31a0581:fenA5B7Qr9FtFjw5@cluster0.9mhf5ll.mongodb.net/test', tlsCAFile=certifi.where())
db = client['contestDetails']
collection = db['users']

# load the workbook
workbook = load_workbook('List.xlsx')

# select the worksheet by name
worksheet = workbook['Worksheet']

# iterate through each row in the worksheet and insert it into MongoDB
for row in worksheet.iter_rows(values_only=True):
    data = {
        'collegeId': row[2],
        'name': row[3],
        'codechefId': row[6]
    }
    collection.insert_one(data)
